"""
Tender Document Extractor — Schedule of Rates
==============================================
Extracts line items from construction tender Schedule of Rates PDFs
into a structured Excel workbook.

Strategy
--------
These PDFs use a broken custom font encoding that defeats text extraction
(pypdf / pdfplumber return CID garbage).  Instead we:
  1. Rasterise each page at 300 DPI with pdf2image
  2. Run Tesseract image_to_data() on the full page to get word bounding boxes
  3. Assign every word to a column (Item / Description / Unit / Qty / Rate / Amount)
     based on its X position relative to the known column-separator pixel positions
  4. Align non-description words to the nearest item row by Y position
  5. Reconstruct each item's description from the Description column words
     that fall between that item's Y and the next item's Y

Column separator positions were calibrated empirically from the sample PDF
header row (see COL_SEPS_FRAC below). Re-calibrate if your PDF differs.

Usage
-----
    python extract_tender.py <input.pdf> [output.xlsx] [--dpi 300] [--quiet]

Output
------
    Excel workbook with:
        • "Schedule of Rates" sheet — one row per line item
        • "Summary" sheet — extraction statistics

    Columns: Project | Schedule No | Page | Page Ref |
             Item | Section | Description | Unit | Qty | Rate | Amount
"""

import sys, re, time, argparse
from pathlib import Path

import pytesseract
from pdf2image import convert_from_path
from PIL import ImageEnhance
import pypdf
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Column separator fractional X positions ─────────────────────────────────
# Measured from page width for the standard A4 SOR table layout.
# Keys are the DIVIDER positions (right edge of each column).
# Re-measure with the calibration snippet at the bottom of this file if needed.
COL_SEPS_FRAC = [0.1407, 0.5377, 0.6183, 0.6989, 0.7997]   # 5 dividers → 6 columns
COL_NAMES     = ["item", "desc", "unit", "qty", "rate", "amount"]

# Table body Y extents (fraction of page height)
TABLE_Y0 = 0.10
TABLE_Y1 = 0.87

# OCR confidence threshold (0–100)
OCR_CONF = 20

# Max Y-distance (pixels at 300 DPI) to associate a unit/qty/rate word with an item row
ROW_SNAP_PX = 130

# Words that are noise / header labels in each column
COL_NOISE = {
    "item":   {"item", "i", "te", "ter", "tem"},
    "unit":   {"unit"},
    "qty":    {"qty", "quantity"},
    "rate":   {"rate"},
    "amount": {"amount", "amount _|", "|"},
}

# Valid item label pattern
ITEM_RE = re.compile(r"^[A-Z]$|^\d{1,3}$")

# Patterns for page-level metadata
SCHEDULE_RE  = re.compile(r"SCHEDULE\s+NO[.\s]+([0-9.]+)", re.IGNORECASE)
PAGE_REF_RE  = re.compile(r"\bS\s*\d+[\./]\d+\b")

# Section heading: ALL CAPS line ≥ 5 chars in the description band
SECTION_RE   = re.compile(r"^[A-Z][A-Z\s/\-&()]{4,}$")

# Noise lines to strip from description
DESC_NOISE_RE = re.compile(
    r"^(SCHEDULE\s+NO\.|JSSC\s+|WORKS\s+AND|BUILDING\s+\d|^Description$|"
    r"TOTAL\s+CARRIED|\(Qn\.|^Notes?[:\-]|^SERVICES$)",
    re.IGNORECASE,
)


# ─── Helpers ─────────────────────────────────────────────────────────────────

def fix_item_label(raw: str) -> str:
    """Normalise OCR noise in item column: 'Cc'→'C', '5)'→'C', etc."""
    t = raw.upper().strip()
    # Double-letter duplicates (Cc, BB): take first char
    if len(t) == 2 and t[0] == t[1]:
        return t[0]
    # Letter + punctuation artifact
    if len(t) == 2 and t[1] in (".", ",", ")", "'", "|"):
        return t[0]
    # Common misreads
    MISREAD = {"5)": "C", "8)": "B", "0)": "O", "|": ""}
    return MISREAD.get(t, t if len(t) == 1 else "")


def is_num_or_dash(s: str) -> bool:
    return bool(re.match(r"^[\d,.\-–]+$", s.strip()))


def is_cid_garbled(text: str) -> bool:
    return text.count("(cid:") > 3


# ─── Page-level metadata ─────────────────────────────────────────────────────

def page_metadata(img) -> tuple[str, str]:
    """Quick low-res OCR of the full page to pick up schedule no & page ref."""
    small = img.resize((img.width // 3, img.height // 3))
    text  = pytesseract.image_to_string(small, config="--psm 6")
    sch   = (m.group(1) if (m := SCHEDULE_RE.search(text)) else "")
    refs  = PAGE_REF_RE.findall(text)
    ref   = refs[-1].replace(" ", "") if refs else ""
    return sch, ref


# ─── Core per-page extractor ─────────────────────────────────────────────────

def process_page(img) -> tuple[str, str, list[dict]]:
    """
    Extract line items from one page image.
    Returns (schedule_no, page_ref, rows).
    """
    img_e = ImageEnhance.Contrast(img).enhance(1.8)
    w, h  = img_e.size

    schedule_no, page_ref = page_metadata(img)

    # --- word bounding boxes ------------------------------------------------
    seps_px = [int(f * w) for f in COL_SEPS_FRAC] + [w]

    def col_of(x: int) -> str:
        for name, sep in zip(COL_NAMES, seps_px):
            if x < sep:
                return name
        return "amount"

    df = pytesseract.image_to_data(img_e,
                                   output_type=pytesseract.Output.DATAFRAME,
                                   config="--psm 6")
    df = df[(df["conf"] > OCR_CONF) & (df["text"].str.strip().str.len() > 0)].copy()

    Y0_PX, Y1_PX = int(TABLE_Y0 * h), int(TABLE_Y1 * h)
    body = df[(df["top"] > Y0_PX) & (df["top"] < Y1_PX)].copy()
    body["col"] = body["left"].apply(col_of)

    # --- item labels ---------------------------------------------------------
    items_raw = body[body["col"] == "item"].copy()
    item_rows: list[dict] = []
    seen_labels: set[str] = set()

    for _, ir in items_raw.sort_values("top").iterrows():
        label = fix_item_label(str(ir["text"]))
        if not ITEM_RE.match(label):
            continue
        if label in seen_labels:          # skip duplicate OCR hits for same cell
            continue
        seen_labels.add(label)
        item_rows.append({"item": label, "item_y": int(ir["top"])})

    if not item_rows:
        return schedule_no, page_ref, []

    # --- assign unit / qty / rate / amount to nearest item row --------------
    item_ys = [r["item_y"] for r in item_rows]

    for col_name in ("unit", "qty", "rate", "amount"):
        col_words = body[body["col"] == col_name].copy()
        noise = COL_NOISE.get(col_name, set())
        col_words = col_words[~col_words["text"].str.lower().str.strip().isin(noise)]

        # For each item row, collect words within ROW_SNAP_PX
        for row in item_rows:
            iy = row["item_y"]
            near = col_words[((col_words["top"] - iy).abs() < ROW_SNAP_PX)]
            tokens = near.sort_values("left")["text"].tolist()
            row[col_name] = " ".join(tokens)

    # --- build description per item row -------------------------------------
    desc_words = body[body["col"] == "desc"].sort_values("top")

    for i, row in enumerate(item_rows):
        iy = row["item_y"]
        y_start = iy - 70
        y_end   = item_ys[i + 1] - 70 if i + 1 < len(item_ys) else Y1_PX

        seg   = desc_words[(desc_words["top"] >= y_start) & (desc_words["top"] < y_end)]
        # Words sorted by (top, left) so they read naturally
        seg_s = seg.sort_values(["top", "left"])
        # Group by row (same top ± 8px) then join left-to-right, rows top-to-bottom
        lines_out = []
        prev_top  = None
        line_buf  = []
        for _, wd in seg_s.iterrows():
            if prev_top is None or abs(wd["top"] - prev_top) > 8:
                if line_buf:
                    lines_out.append(" ".join(line_buf))
                line_buf  = [wd["text"]]
                prev_top  = wd["top"]
            else:
                line_buf.append(wd["text"])
        if line_buf:
            lines_out.append(" ".join(line_buf))

        # Filter noise lines
        lines_out = [l for l in lines_out if not DESC_NOISE_RE.match(l)]
        desc = " ".join(lines_out).strip()
        desc = re.sub(r"\s*\(?Cont'?d\)?", "", desc, flags=re.IGNORECASE).strip()
        row["desc"] = desc

    # --- detect section heading ---------------------------------------------
    all_desc_lines = [
        " ".join(desc_words[
            (desc_words["top"] >= int(TABLE_Y0 * h)) &
            (desc_words["top"] < (item_ys[0] - 70 if item_ys else Y1_PX))
        ].sort_values("left")["text"].tolist())
    ]
    section = ""
    for line in all_desc_lines:
        if SECTION_RE.match(line.strip()):
            section = line.strip()
            break

    # --- clean up and return ------------------------------------------------
    rows = []
    for row in item_rows:
        unit = row.get("unit", "").lower().strip()
        qty  = row.get("qty",  "").strip()
        rate = row.get("rate", "").strip()
        amt  = row.get("amount", "").strip()

        rows.append({
            "Item":        row["item"],
            "Section":     section,
            "Description": row.get("desc", ""),
            "Unit":        unit  if re.match(r"^(item|nr|no|m2|m3|m|ls|lsum|set|lot|run|%|pc|pcs|roll|\-)$", unit) else "",
            "Qty":         qty   if is_num_or_dash(qty)  else "",
            "Rate":        rate  if is_num_or_dash(rate) else "",
            "Amount":      amt   if is_num_or_dash(amt)  else "",
        })

    return schedule_no, page_ref, rows


# ─── PDF orchestrator ────────────────────────────────────────────────────────

def get_project_name(pdf_path: str) -> str:
    try:
        text = pypdf.PdfReader(pdf_path).pages[0].extract_text() or ""
        if is_cid_garbled(text):
            return ""
        for line in text.split("\n"):
            line = line.strip()
            # Only return lines that are mostly printable ASCII
            printable = sum(32 <= ord(c) < 127 for c in line)
            if len(line) > 15 and printable / max(len(line), 1) > 0.85:
                return line
    except Exception:
        pass
    return ""


def process_pdf(pdf_path: str, dpi: int = 300, verbose: bool = True) -> pd.DataFrame:
    reader      = pypdf.PdfReader(pdf_path)
    total_pages = len(reader.pages)
    project     = get_project_name(pdf_path)

    if verbose:
        print(f"\nTender Extractor  v2.0")
        print(f"{'─'*55}")
        print(f"Input   : {pdf_path}")
        print(f"Pages   : {total_pages}  |  DPI: {dpi}")
        if project:
            print(f"Project : {project}")
        print(f"{'─'*55}")

    t0, all_rows = time.time(), []

    for pg in range(total_pages):
        pn = pg + 1
        if verbose:
            print(f"  Page {pn:>2}/{total_pages}  [{time.time()-t0:5.1f}s]  ", end="", flush=True)
        try:
            images = convert_from_path(pdf_path, dpi=dpi, first_page=pn, last_page=pn)
            if not images:
                if verbose: print("(no image)")
                continue

            sch, ref, rows = process_page(images[0])

            for row in rows:
                row.update({"Project": project, "Schedule No": sch,
                             "Page": pn,         "Page Ref":   ref})
                all_rows.append(row)

            if verbose:
                print(f"{len(rows)} items  (Sch {sch or '?'}, {ref or '?'})")
        except Exception as exc:
            if verbose:
                print(f"ERROR — {exc}")

    df = pd.DataFrame(all_rows, columns=[
        "Project", "Schedule No", "Page", "Page Ref",
        "Item", "Section", "Description", "Unit", "Qty", "Rate", "Amount"
    ])
    if verbose:
        print(f"{'─'*55}")
        print(f"Done    : {len(df)} rows extracted in {time.time()-t0:.1f}s")
    return df


# ─── Excel writer ────────────────────────────────────────────────────────────

DARK_BLUE = "1F3864"
MID_BLUE  = "2E75B6"
ALT_ROW   = "EBF3FB"
_TS = Side(style="thin", color="BFBFBF")
TB  = Border(left=_TS, right=_TS, top=_TS, bottom=_TS)


def _c(ws, r, c, v, bold=False, fg="000000", bg=None, ha="left", wrap=False, sz=9):
    cell = ws.cell(r, c, v)
    cell.font      = Font(bold=bold, color=fg, size=sz)
    cell.alignment = Alignment(horizontal=ha, vertical="top", wrap_text=wrap)
    cell.border    = TB
    if bg:
        cell.fill  = PatternFill("solid", fgColor=bg)
    return cell


def save_to_excel(df: pd.DataFrame, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule of Rates"

    COLS = [
        ("Project",     36, "left"),
        ("Schedule No", 13, "center"),
        ("Page",         6, "center"),
        ("Page Ref",    10, "center"),
        ("Item",         6, "center"),
        ("Section",     28, "left"),
        ("Description", 56, "left"),
        ("Unit",         8, "center"),
        ("Qty",          8, "center"),
        ("Rate",        13, "right"),
        ("Amount",      13, "right"),
    ]

    for ci, (name, w, _) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Header row
    ws.row_dimensions[1].height = 30
    for ci, (name, _, _) in enumerate(COLS, 1):
        _c(ws, 1, ci, name, bold=True, fg="FFFFFF", bg=DARK_BLUE, ha="center", sz=10)

    # Data rows
    dr, prev_sch = 2, None
    for _, rec in df.iterrows():
        sch = rec.get("Schedule No", "")

        # Schedule banner row
        if sch and sch != prev_sch:
            ws.merge_cells(start_row=dr, start_column=1, end_row=dr, end_column=len(COLS))
            banner = ws.cell(dr, 1, f"  ■  SCHEDULE NO. {sch}")
            banner.font      = Font(bold=True, color="FFFFFF", size=10)
            banner.fill      = PatternFill("solid", fgColor=MID_BLUE)
            banner.alignment = Alignment(horizontal="left", vertical="center")
            banner.border    = TB
            ws.row_dimensions[dr].height = 22
            dr += 1
            prev_sch = sch

        bg = ALT_ROW if dr % 2 == 0 else "FFFFFF"
        vals = [rec.get(col[0], "") for col in COLS]
        for ci, (v, (_, _, ha)) in enumerate(zip(vals, COLS), 1):
            _c(ws, dr, ci, v, bg=bg, ha=ha, wrap=(ci == 7))
        ws.row_dimensions[dr].height = 28
        dr += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 24
    summary = [
        ("Field",            "Value"),
        ("Total Line Items",  len(df)),
        ("Schedules Found",   ", ".join(df["Schedule No"].dropna().unique()) or "N/A"),
        ("Pages with Items",  int(df["Page"].nunique())),
        ("Items with Qty",
         int((df["Qty"].str.strip().replace({"-": "", "–": ""}) != "").sum())),
        ("Tool",             "extract_tender.py v2.0"),
    ]
    for r, (label, val) in enumerate(summary, 1):
        ws2.cell(r, 1, label).font = Font(bold=(r == 1),
                                          color="FFFFFF" if r == 1 else "000000", size=10)
        ws2.cell(r, 2, str(val)).font = Font(size=10)
        if r == 1:
            for c in (ws2.cell(r, 1), ws2.cell(r, 2)):
                c.fill      = PatternFill("solid", fgColor=DARK_BLUE)
                c.alignment = Alignment(horizontal="center")

    wb.save(output_path)
    print(f"\n✓  Saved → {output_path}   ({len(df)} rows, "
          f"{df['Schedule No'].nunique()} schedule(s))")


# ─── CLI ─────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description="Extract Schedule of Rates from construction tender PDFs"
    )
    ap.add_argument("pdf",            help="Input PDF")
    ap.add_argument("output", nargs="?",
                    help="Output .xlsx  (default: <pdf_stem>_extracted.xlsx)")
    ap.add_argument("--dpi",      type=int,   default=300,
                    help="Rasterisation DPI (default 300; 200 = faster but less accurate)")
    ap.add_argument("--quiet",    action="store_true", help="Suppress progress output")
    ap.add_argument("--no-excel", action="store_true", help="Print CSV to stdout instead")
    args = ap.parse_args()

    if not Path(args.pdf).exists():
        sys.exit(f"Error: {args.pdf} not found")

    out = args.output or (Path(args.pdf).stem + "_extracted.xlsx")
    df  = process_pdf(args.pdf, dpi=args.dpi, verbose=not args.quiet)

    if df.empty:
        print("\n⚠  No line items extracted.")
        print("   If your PDF uses different column proportions, adjust COL_SEPS_FRAC")
        print("   at the top of this script using the calibration snippet below.\n")
        print("   Calibration snippet (run in Python after rasterising page 1 at 300 DPI):")
        print("     import numpy as np")
        print("     arr = np.array(img.convert('L'))")
        print("     # scan header row y=300-380 for dark vertical lines")
        print("     row = arr[300:380, :]")
        print("     dark = np.where(row.min(axis=0) < 50)[0]")
        print("     print('separator px:', dark[::5])   # cluster manually")
        sys.exit(0)

    if args.no_excel:
        print(df.to_csv(index=False))
    else:
        save_to_excel(df, out)


if __name__ == "__main__":
    main()


# ─────────────────────────────────────────────────────────────────────────────
# CALIBRATION SNIPPET
# Run this block once with your PDF to find COL_SEPS_FRAC for a new layout.
# ─────────────────────────────────────────────────────────────────────────────
# import numpy as np
# from pdf2image import convert_from_path
# imgs = convert_from_path("your.pdf", dpi=300, first_page=1, last_page=1)
# img  = imgs[0]
# w, h = img.size
# arr  = np.array(img.convert("L"))
# # scan the header band for dark vertical lines
# header_band = arr[int(0.09*h):int(0.12*h), :]
# col_min = header_band.min(axis=0)
# dark_cols = [i for i in range(5, w-5) if col_min[i] < 50]
# # cluster contiguous pixels into separator centres
# clusters, cl = [], [dark_cols[0]]
# for p in dark_cols[1:]:
#     if p - cl[-1] <= 5: cl.append(p)
#     else:
#         clusters.append(int(sum(cl)/len(cl))); cl = [p]
# clusters.append(int(sum(cl)/len(cl)))
# print("COL_SEPS_FRAC =", [round(c/w, 4) for c in clusters])
